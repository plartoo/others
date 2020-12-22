"""This is the subclass of Transform function for Netherlands (EU division).

We will define transform functions specific to Netherlands here.

Author: Maicol Contreras
Last Modified: December 21, 2020
"""
from constants import comp_harm_constants
import pandas as pd
from openpyxl import load_workbook
from transform_functions.common_comp_harm_transform_functions import CommonCompHarmTransformFunctions


class EuNetherlandsTransformFunctions(CommonCompHarmTransformFunctions):
    """
    All custom (uncommon) transform functions **SPECIFIC to
    individual processing task** must be defined as part
    of this class.
    """
    NETHERLANDS_SPECIFIC_CATEGORY_MAPPINGS = {
        # Add any Netherlands-specific mappings that cannot be used for other EU countries here.
    }

    def __init__(self, config):
        self.config = config
        self.category_mappings = dict(
            comp_harm_constants.ENGLISH_CATEGORY_MAPPINGS,
            **EuNetherlandsTransformFunctions.NETHERLANDS_SPECIFIC_CATEGORY_MAPPINGS)

    def join_sheets_in_a_unique_dataframse(self,df):
        columns_to_use = ['Productklasse', 'Merk', 'Product', 'Adverteerder',
        'Mediumtype', 'Jaar', 'Maand', 'Week', 'Spend']
        final_df = pd.DataFrame(columns=columns_to_use)
        wb = load_workbook(self.config['current_input_file'])
        worksheets = wb.sheetnames
        worksheets_to_use = [sheet for sheet in worksheets if len(sheet.split())==2 and sheet.split()[1].isnumeric()]
        for worksheet in worksheets_to_use:
            dt_to_clean = pd.read_excel(self.config['current_input_file'],skiprows=self.config['header'],sheet_name=worksheet)
            dt_to_clean.dropna(thresh=9,inplace=True)
            dt_to_clean['Jaar'] = dt_to_clean['Jaar'].astype(int)
            dt_to_clean = dt_to_clean[~dt_to_clean['Spend'].isnull()]
            final_df = final_df.append(dt_to_clean[columns_to_use],sort=False)
        return final_df